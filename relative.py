#!/usr/bin/env python3
"""
Script to add relative time to all Excel files based on EDF metadata start time

Features:
1. Set reference time from EDF metadata start time
2. Find all Excel files of the same date (YYYYMMDD)
3. Calculate and add relative time to each file using EDF metadata as reference
"""

import os
import glob
import re
import openpyxl
from datetime import datetime
import mne

def time_to_seconds(time_str):
    """Convert time string in HH:MM:SS.SS format to seconds"""
    if time_str is None or time_str == '':
        return 0
    
    # Clean string (remove leading/trailing spaces)
    time_str = str(time_str).strip()
    
    # Check if it's in HH:MM:SS.SS format
    time_pattern = r'(\d{1,2}):(\d{2}):(\d{2}\.?\d*)'
    match = re.match(time_pattern, time_str)
    
    if match:
        hours = int(match.group(1))
        minutes = int(match.group(2))
        seconds = float(match.group(3))
        
        total_seconds = hours * 3600 + minutes * 60 + seconds
        return total_seconds
    else:
        print(f"  Warning: Cannot recognize time format: '{time_str}'")
        return 0

def parse_edf_time_to_seconds(edf_time_str):
    """Parse EDF time string (HH.MM.SS) to seconds"""
    try:
        # EDF time format: HH.MM.SS (with dots)
        parts = edf_time_str.split('.')
        if len(parts) >= 3:
            hours = float(parts[0])
            minutes = float(parts[1])
            seconds = float(parts[2])
            return hours * 3600 + minutes * 60 + seconds
        elif len(parts) == 2:
            # If only HH.MM format, assume seconds as 0
            hours = float(parts[0])
            minutes = float(parts[1])
            seconds = 0.0
            return hours * 3600 + minutes * 60 + seconds
        return 0.0
    except:
        return 0.0

def extract_edf_metadata(edf_file):
    """Extract metadata from EDF file header"""
    try:
        with open(edf_file, 'rb') as f:
            # Read EDF header (256 bytes)
            header_bytes = f.read(256)
            header_str = header_bytes.decode('ascii', errors='ignore')
            
            # Parse header information
            metadata = {}
            metadata['version'] = header_str[0:8].strip()
            metadata['patient_id'] = header_str[8:88].strip()
            metadata['recording_id'] = header_str[88:168].strip()
            metadata['start_date'] = header_str[168:176].strip()  # DD.MM.YY format
            metadata['start_time'] = header_str[176:184].strip()  # HH.MM.SS format
            metadata['header_bytes'] = header_str[184:192].strip()
            metadata['reserved'] = header_str[192:236].strip()
            metadata['records'] = header_str[236:244].strip()
            metadata['duration'] = header_str[244:252].strip()
            metadata['nsignals'] = header_str[252:256].strip()
            
            return metadata
    except Exception as e:
        print(f"  Error reading EDF metadata: {e}")
        return None

def get_edf_reference_time(edf_file):
    """Get reference time from EDF metadata start time"""
    try:
        metadata = extract_edf_metadata(edf_file)
        if not metadata:
            return None
        
        start_time_str = metadata['start_time']
        reference_seconds = parse_edf_time_to_seconds(start_time_str)
        
        print(f"  EDF start time: {start_time_str}")
        print(f"  Reference time (EDF start): {reference_seconds:.2f}s")
        return reference_seconds
        
    except Exception as e:
        print(f"Error getting EDF reference time: {e}")
        return None

def find_matching_edf_file(excel_file):
    """Find matching EDF file for the given Excel file"""
    try:
        # Extract base name from Excel file (remove .xlsx)
        base_name = os.path.splitext(excel_file)[0]
        
        # Look for EDF file with same base name
        edf_file = base_name + '.edf'
        if os.path.exists(edf_file):
            return edf_file
        
        # If not found, try to find in the same directory
        directory = os.path.dirname(excel_file)
        if directory:
            edf_pattern = os.path.join(directory, os.path.basename(base_name) + '*.edf')
            edf_files = glob.glob(edf_pattern)
            if edf_files:
                return edf_files[0]  # Return first match
        
        return None
        
    except Exception as e:
        print(f"  Error finding EDF file for {excel_file}: {e}")
        return None

def process_excel_file(file_path, start_time):
    """Process single Excel file to add relative time"""
    print(f"\nProcessing: {file_path}")
    
    try:
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        # Convert data to list
        data = []
        for row_num in range(1, ws.max_row + 1):
            row_values = []
            for col_num in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=row_num, column=col_num).value
                row_values.append(cell_value)
            data.append(row_values)
        
        # Remove first empty row
        if data and all(cell is None for cell in data[0]):
            data = data[1:]
            print("  Removed first empty row.")
        
        print(f"  Data size: {len(data)} rows x {len(data[0]) if data else 0} columns")
        
        # Find time column (column C, index 2)
        time_col = 2
        if len(data[0]) < 3:
            print("  Warning: Column C not found.")
            return False
        
        # Add column E (index 4) and F (index 5) for relative time and status
        for i in range(len(data)):
            while len(data[i]) < 6:
                data[i].append(None)
        
        # Convert each row's time to relative time and store in column E
        # Initialize status column F with "PENDING" (will be updated by edf2edfplus.py)
        for i in range(len(data)):
            if len(data[i]) > time_col:
                time_val = data[i][time_col]
                if time_val is not None and str(time_val).strip() != '':
                    current_time = time_to_seconds(time_val)
                    relative_time = current_time - start_time
                    data[i][4] = f"{relative_time:.2f}"  # Column E (index 4)
                    data[i][5] = "PENDING"  # Column F (index 5) - Status
                else:
                    data[i][4] = ''
                    data[i][5] = ''
            else:
                data[i][4] = ''
                data[i][5] = ''
        
        # Check results
        print("  Conversion results:")
        for i in range(min(3, len(data))):  # Show first 3 rows only
            original_time = data[i][time_col] if len(data[i]) > time_col else None
            relative_time = data[i][4] if len(data[i]) > 4 else None
            status = data[i][5] if len(data[i]) > 5 else None
            print(f"    Row {i+1}: {original_time} -> {relative_time}s [{status}]")
        
        # Create new workbook and save
        new_wb = openpyxl.Workbook()
        new_ws = new_wb.active
        
        # Write data to new worksheet
        for i, row_data in enumerate(data):
            for j, cell_value in enumerate(row_data):
                if cell_value is not None:
                    new_ws.cell(row=i+1, column=j+1, value=cell_value)
        
        # Save file
        new_wb.save(file_path)
        print(f"  Save complete: {file_path}")
        
        return True
        
    except Exception as e:
        print(f"  Error occurred: {e}")
        return False

def get_edf_reference_times(file_list):
    """Get EDF reference times for all Excel files"""
    reference_times = {}
    
    for file_path in file_list:
        try:
            # Find matching EDF file
            edf_file = find_matching_edf_file(file_path)
            if edf_file:
                reference_time = get_edf_reference_time(edf_file)
                if reference_time is not None:
                    reference_times[file_path] = reference_time
                    print(f"  Found EDF reference for {os.path.basename(file_path)}: {reference_time:.2f}s")
                else:
                    print(f"  Warning: Could not get EDF reference for {os.path.basename(file_path)}")
            else:
                print(f"  Warning: No matching EDF file found for {os.path.basename(file_path)}")
        
        except Exception as e:
            print(f"  Error processing {file_path}: {e}")
            continue
    
    return reference_times

def main():
    """Main function"""
    print("=== Starting Excel event file time conversion (EDF-based) ===")
    
    # Find all Excel files with date pattern
    excel_pattern = "**/*_*_*.xlsx"  # Pattern: anything_YYYYMMDD_HHMM.xlsx
    all_excel_files = glob.glob(excel_pattern, recursive=True)
    all_excel_files = [f for f in all_excel_files if not os.path.basename(f).startswith('~$')]
    
    if not all_excel_files:
        print("No Excel files found.")
        return
    
    print(f"Found Excel files: {len(all_excel_files)}")
    
    # Group files by date
    date_groups = {}
    for file_path in all_excel_files:
        # Extract date from filename (YYYYMMDD)
        date_match = re.search(r'(\d{8})', file_path)
        if date_match:
            date_str = date_match.group(1)
            if date_str not in date_groups:
                date_groups[date_str] = []
            date_groups[date_str].append(file_path)
    
    print(f"Found date groups: {len(date_groups)}")
    
    # Process each date group
    total_success = 0
    total_files = 0
    
    for date_str, file_list in date_groups.items():
        print(f"\n=== Processing date {date_str} ===")
        print(f"  Files: {len(file_list)}")
        for file in file_list:
            print(f"    - {os.path.basename(file)}")
        
        # Get EDF reference times for all files in this date group
        reference_times = get_edf_reference_times(file_list)
        if not reference_times:
            print(f"  Cannot find EDF reference times for files in date {date_str}")
            continue
        
        print(f"  Found EDF references for {len(reference_times)} files")
        
        # Process each file with its own EDF reference time
        success_count = 0
        for file_path in file_list:
            if file_path in reference_times:
                if process_excel_file(file_path, reference_times[file_path]):
                    success_count += 1
            else:
                print(f"  Skipping {os.path.basename(file_path)} - no EDF reference time")
            total_files += 1
        
        total_success += success_count
        print(f"  Date {date_str} complete: {success_count}/{len(file_list)} success")
    
    print(f"\n=== Overall processing complete ===")
    print(f"Total success: {total_success}/{total_files} files")

if __name__ == "__main__":
    main()
