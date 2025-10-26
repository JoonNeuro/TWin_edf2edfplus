#!/usr/bin/env python3
"""
Rollback script to restore original EDF files and remove EDF+ files

This script will:
1. Find all _backup.edf files
2. Restore them to original names (remove _backup suffix)
3. Remove the EDF+ files that were created
4. Remove relative time data (Column E) from Excel files
"""

import os
import glob
from pathlib import Path
import shutil
import openpyxl
import logging
from datetime import datetime

def setup_logging():
    """Setup logging to both console and file"""
    # Create log filename with current timestamp
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    log_filename = f"log_rollback_{timestamp}.log"
    
    # Configure logging
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler(log_filename, encoding='utf-8'),
            logging.StreamHandler()
        ]
    )
    
    return log_filename

def log_print(message):
    """Print message to both console and log file"""
    print(message)
    logging.info(message)

def rollback_excel_files():
    """Remove relative time data (Column E) from Excel files"""
    log_print("\n=== Excel File Rollback ===")
    
    # Find all Excel files
    excel_files = glob.glob("**/*.xlsx", recursive=True)
    excel_files = [f for f in excel_files if not os.path.basename(f).startswith('~$')]
    
    if not excel_files:
        log_print("No Excel files found.")
        return
    
    log_print(f"Found {len(excel_files)} Excel files:")
    for excel_file in excel_files:
        log_print(f"  - {excel_file}")
    
    # Process each Excel file
    processed_count = 0
    
    for excel_file in excel_files:
        log_print(f"\nProcessing: {excel_file}")
        
        try:
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active
            
            # Convert data to list
            data = []
            for row_num in range(1, ws.max_row + 1):
                row_values = []
                for col_num in range(1, ws.max_column + 1):
                    cell_value = ws.cell(row=row_num, column=col_num).value
                    row_values.append(cell_value)
                data.append(row_values)
            
            # Check if Column E (index 4) exists and has data
            has_relative_time = False
            for i in range(len(data)):
                if len(data[i]) > 4 and data[i][4] is not None and str(data[i][4]).strip() != '':
                    has_relative_time = True
                    break
            
            if not has_relative_time:
                log_print("  - No relative time data found, skipping")
                continue
            
            # Remove Column E (index 4) data
            for i in range(len(data)):
                if len(data[i]) > 4:
                    data[i][4] = None  # Clear Column E
            
            # Create new workbook and save
            new_wb = openpyxl.Workbook()
            new_ws = new_wb.active
            
            # Write data to new worksheet (without Column E)
            for i, row_data in enumerate(data):
                for j, cell_value in enumerate(row_data):
                    if cell_value is not None and j != 4:  # Skip Column E
                        new_ws.cell(row=i+1, column=j+1, value=cell_value)
            
            # Save file
            new_wb.save(excel_file)
            log_print(f"  ✓ Removed relative time data from Column E")
            processed_count += 1
            
        except Exception as e:
            log_print(f"  ✗ Error processing {excel_file}: {e}")
    
    log_print(f"\n=== Excel Rollback Complete ===")
    log_print(f"Processed files: {processed_count}")

def rollback_edf_files():
    """Rollback EDF files to original state"""
    log_print("=== EDF File Rollback ===")
    
    # Find all _backup_*.edf files (new naming pattern)
    backup_files = glob.glob("**/*_backup_*.edf", recursive=True)
    
    if not backup_files:
        log_print("No backup files found.")
        return
    
    log_print(f"Found {len(backup_files)} backup files:")
    for backup_file in backup_files:
        log_print(f"  - {backup_file}")
    
    # Process each backup file
    restored_count = 0
    removed_edfplus_count = 0
    
    for backup_file in backup_files:
        backup_path = Path(backup_file)
        
        # Extract original filename from backup filename
        # Pattern: originalname_backup_finalname.edf -> originalname.edf
        backup_stem = backup_path.stem
        if "_backup_" in backup_stem:
            original_name = backup_stem.split("_backup_")[0] + ".edf"
        else:
            # Fallback for old naming pattern
            original_name = backup_stem.replace("_backup", "") + ".edf"
        
        original_path = backup_path.parent / original_name
        
        # Find the EDF+ file (final filename)
        # Pattern: originalname_backup_finalname.edf -> finalname.edf
        if "_backup_" in backup_stem:
            final_name = backup_stem.split("_backup_")[1] + ".edf"
            edfplus_path = backup_path.parent / final_name
        else:
            edfplus_path = None
        
        log_print(f"\nProcessing: {backup_path.name}")
        log_print(f"  Original: {original_name}")
        if edfplus_path:
            log_print(f"  EDF+ file: {final_name}")
        
        try:
            # Restore backup to original name
            shutil.move(str(backup_path), str(original_path))
            log_print(f"  ✓ Restored backup: {original_name}")
            restored_count += 1
            
            # Remove the EDF+ file if it exists
            if edfplus_path and edfplus_path.exists():
                edfplus_path.unlink()
                log_print(f"  ✓ Removed EDF+ file: {final_name}")
                removed_edfplus_count += 1
            
        except Exception as e:
            log_print(f"  ✗ Error processing {backup_path.name}: {e}")
    
    log_print(f"\n=== Rollback Complete ===")
    log_print(f"Restored files: {restored_count}")
    log_print(f"Removed EDF+ files: {removed_edfplus_count}")

def main():
    """Main function"""
    # Setup logging
    log_filename = setup_logging()
    log_print("=== Rollback Started ===")
    log_print(f"Log file: {log_filename}")
    
    log_print("Starting complete rollback...")
    
    # Confirm before proceeding
    response = input("This will restore original EDF files, remove EDF+ files, and remove relative time data from Excel files. Continue? (y/N): ")
    if response.lower() != 'y':
        log_print("Rollback cancelled.")
        return
    
    # Rollback EDF files
    rollback_edf_files()
    
    # Rollback Excel files
    rollback_excel_files()
    
    log_print("\n=== Complete Rollback Finished ===")
    log_print("All changes have been reverted to original state.")
    log_print(f"Log saved to: {log_filename}")

if __name__ == "__main__":
    main()
