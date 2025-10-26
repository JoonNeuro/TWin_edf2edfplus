#!/usr/bin/env python3
"""
EDF to EDF+ Converter with Excel Events
MNE Python implementation for merging EDF files with Excel event files and saving as EDF+ format
"""

import mne
import pandas as pd
import numpy as np
from datetime import datetime
import os
import sys
from pathlib import Path
import glob
import re
import struct
import logging


def setup_logging():
    """Setup logging to both console and file"""
    # Create log filename with current timestamp
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    log_filename = f"log_edf2edfplus_{timestamp}.log"
    
    # Configure logging
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler(log_filename, encoding='utf-8'),
            logging.StreamHandler(sys.stdout)
        ]
    )
    
    # Suppress MNE warnings in log file but show in console
    mne.set_log_level('WARNING')
    
    return log_filename


def log_print(message):
    """Print message to both console and log file"""
    print(message)
    logging.info(message)


def parse_time_to_seconds(time_str):
    """Parse time string (HH:MM:SS.ss) to seconds"""
    try:
        parts = time_str.split(':')
        if len(parts) >= 3:
            hours = float(parts[0])
            minutes = float(parts[1])
            seconds = float(parts[2])
            return hours * 3600 + minutes * 60 + seconds
        return 0.0
    except:
        return 0.0


def parse_edf_time_to_seconds(edf_time_str):
    """Parse EDF time string (HH.MM.SS) to seconds"""
    try:
        # EDF time format: HH.MM.SS (with dots)
        parts = edf_time_str.split('.')
        if len(parts) >= 3:
            hours = float(parts[0])
            minutes = float(parts[1])
            seconds = float(parts[2])
            return hours * 3600 + minutes * 60 + seconds
        elif len(parts) == 2:
            # If only HH.MM format, assume seconds as 0
            hours = float(parts[0])
            minutes = float(parts[1])
            seconds = 0.0
            return hours * 3600 + minutes * 60 + seconds
        return 0.0
    except:
        return 0.0


def extract_edf_metadata(edf_file):
    """Extract metadata from EDF file header"""
    try:
        with open(edf_file, 'rb') as f:
            # Read EDF header (256 bytes)
            header_bytes = f.read(256)
            header_str = header_bytes.decode('ascii', errors='ignore')
            
            # Parse header information
            metadata = {}
            metadata['version'] = header_str[0:8].strip()
            metadata['patient_id'] = header_str[8:88].strip()
            metadata['recording_id'] = header_str[88:168].strip()
            metadata['start_date'] = header_str[168:176].strip()  # DD.MM.YY format
            metadata['start_time'] = header_str[176:184].strip()  # HH.MM.SS format
            metadata['header_bytes'] = header_str[184:192].strip()
            metadata['reserved'] = header_str[192:236].strip()
            metadata['records'] = header_str[236:244].strip()
            metadata['duration'] = header_str[244:252].strip()
            metadata['nsignals'] = header_str[252:256].strip()
            
            return metadata
    except Exception as e:
        log_print(f"Error reading EDF metadata: {e}")
        return None


def convert_edf_date_time(date_str, time_str):
    """Convert EDF date/time format to YYYYMMDD_HHMM format"""
    try:
        # Parse date: DD.MM.YY -> YYYYMMDD
        day, month, year = date_str.split('.')
        full_year = 2000 + int(year) if int(year) < 50 else 1900 + int(year)
        date_formatted = f"{full_year:04d}{int(month):02d}{int(day):02d}"
        
        # Parse time: HH.MM.SS -> HHMM
        hour, minute, second = time_str.split('.')
        time_formatted = f"{int(hour):02d}{int(minute):02d}"
        
        return f"{date_formatted}_{time_formatted}"
    except Exception as e:
        log_print(f"Error converting date/time: {e}")
        return None


def generate_correct_filename(edf_file):
    """Generate correct filename based on EDF metadata"""
    try:
        # Extract metadata
        metadata = extract_edf_metadata(edf_file)
        if not metadata:
            return None
        
        # Get patient ID and convert date/time
        patient_id = metadata['patient_id'].strip()
        if not patient_id or patient_id == "No Database Record":
            # Extract patient ID from filename
            import re
            match = re.search(r'(\d+)_', os.path.basename(edf_file))
            if match:
                patient_id = match.group(1)
            else:
                log_print(f"Could not extract patient ID from filename: {edf_file}")
                return None
        
        # Convert EDF date/time to YYYYMMDD_HHMM format
        date_time_str = convert_edf_date_time(metadata['start_date'], metadata['start_time'])
        if not date_time_str:
            return None
        
        # Generate new filename
        new_filename = f"{patient_id}_{date_time_str}.edf"
        
        log_print(f"  Original filename: {os.path.basename(edf_file)}")
        log_print(f"  EDF metadata - Date: {metadata['start_date']}, Time: {metadata['start_time']}")
        log_print(f"  Generated filename: {new_filename}")
        
        return new_filename
        
    except Exception as e:
        log_print(f"Error generating filename: {e}")
        return None


def read_edf_file_direct(filename):
    """Read EDF file directly (like MATLAB)"""
    with open(filename, 'rb') as f:
        # Read EDF header (256 bytes)
        header_bytes = f.read(256)
        header_str = header_bytes.decode('ascii', errors='ignore')
        
        # Parse header information
        header = {}
        try:
            header['version'] = float(header_str[0:8].strip() or '0')
        except:
            header['version'] = 0.0
        header['patient_id'] = header_str[8:88].strip()
        header['recording_id'] = header_str[88:168].strip()
        header['start_date'] = header_str[168:176].strip()
        header['start_time'] = header_str[176:184].strip()
        try:
            header['header_bytes'] = int(header_str[184:192].strip() or '0')
        except:
            header['header_bytes'] = 0
        header['reserved'] = header_str[192:236].strip()
        try:
            header['records'] = int(header_str[236:244].strip() or '0')
        except:
            header['records'] = 0
        try:
            header['duration'] = float(header_str[244:252].strip() or '0')
        except:
            header['duration'] = 0.0
        try:
            header['nsignals'] = int(header_str[252:256].strip() or '0')
        except:
            header['nsignals'] = 0
        
        # Read signal headers (nsignals * 256 bytes)
        signal_header_bytes = f.read(header['nsignals'] * 256)
        signal_header_str = signal_header_bytes.decode('ascii', errors='ignore')
        
        # Parse each signal's information
        header['signals'] = []
        for i in range(header['nsignals']):
            start_idx = i * 256
            end_idx = (i + 1) * 256
            signal_str = signal_header_str[start_idx:end_idx]
            
            signal_info = {}
            signal_info['label'] = signal_str[0:16].strip()
            signal_info['transducer'] = signal_str[16:32].strip()
            signal_info['units'] = signal_str[32:40].strip()
            try:
                signal_info['physical_min'] = float(signal_str[40:48].strip() or '0')
            except:
                signal_info['physical_min'] = 0.0
            try:
                signal_info['physical_max'] = float(signal_str[48:56].strip() or '0')
            except:
                signal_info['physical_max'] = 0.0
            try:
                signal_info['digital_min'] = int(signal_str[56:64].strip() or '0')
            except:
                signal_info['digital_min'] = 0
            try:
                signal_info['digital_max'] = int(signal_str[64:72].strip() or '0')
            except:
                signal_info['digital_max'] = 0
            signal_info['prefilter'] = signal_str[72:80].strip()
            try:
                signal_info['samples_per_record'] = int(signal_str[80:88].strip() or '0')
            except:
                signal_info['samples_per_record'] = 0
            signal_info['reserved'] = signal_str[88:256].strip()
            
            header['signals'].append(signal_info)
        
        # Read data
        data = []
        
        # Calculate total samples for each signal (each signal can have different sample counts)
        total_samples_per_signal = []
        for sig in range(header['nsignals']):
            samples_per_record = header['signals'][sig]['samples_per_record']
            total_samples = header['records'] * samples_per_record
            total_samples_per_signal.append(total_samples)
        
        log_print(f"  EDF header info:")
        log_print(f"    Records: {header['records']}")
        log_print(f"    Duration: {header['duration']}")
        log_print(f"    Signals: {header['nsignals']}")
        log_print(f"    Samples per record for each signal: {[s['samples_per_record'] for s in header['signals'][:5]]}")
        
        # Pre-allocate data array (use maximum samples for all signals)
        max_samples = max(total_samples_per_signal) if total_samples_per_signal else 0
        data = np.zeros((header['nsignals'], max_samples), dtype=np.int16)
        
        # Read data for each record
        for rec in range(header['records']):
            for sig in range(header['nsignals']):
                samples = header['signals'][sig]['samples_per_record']
                if samples > 0:
                    signal_data = struct.unpack('<' + 'h' * samples, f.read(samples * 2))
                    if len(signal_data) == samples:
                        start_idx = rec * samples
                        end_idx = (rec + 1) * samples
                        if end_idx <= max_samples:
                            data[sig, start_idx:end_idx] = signal_data
                        else:
                            log_print(f"  Warning: Signal {sig} data exceeds allocated space")
                    else:
                        log_print(f"  Warning: Sample count mismatch in signal {sig}, record {rec}")
                else:
                    # Skip reading for signals with 0 samples per record
                    pass
        
        return header, data


def create_mne_raw_from_edf_data(edf_header, edf_data):
    """Create MNE Raw object from direct EDF data"""
    # Calculate sampling rate
    sampling_rate = edf_header['signals'][0]['samples_per_record'] / edf_header['duration']
    
    # Create channel names
    ch_names = [signal['label'] for signal in edf_header['signals']]
    
    # Create info object
    info = mne.create_info(
        ch_names=ch_names,
        sfreq=sampling_rate,
        ch_types=['eeg'] * len(ch_names)
    )
    
    # Convert int16 to float64 and apply scaling
    data_float = edf_data.astype(np.float64)
    
    # Apply physical scaling
    for i, signal in enumerate(edf_header['signals']):
        if signal['physical_max'] != signal['physical_min']:
            scale = (signal['physical_max'] - signal['physical_min']) / (signal['digital_max'] - signal['digital_min'])
            offset = signal['physical_min'] - scale * signal['digital_min']
            data_float[i, :] = data_float[i, :] * scale + offset
    
    # Create Raw object
    raw = mne.io.RawArray(data_float, info, verbose=False)
    
    return raw


def find_matching_excel_files(edf_file):
    """Find EDF file and matching Excel files"""
    edf_path = Path(edf_file)
    edf_name = edf_path.stem
    
    # Extract patient_number_YYYYMMDD_HHMM pattern from EDF filename
    # Example: 5774131_20130701_2359 -> 5774131_20130701_2359
    # Patient number is one or more digits
    pattern = r'(\d+_\d{8}_\d{4})'
    match = re.search(pattern, edf_name)
    
    if not match:
        log_print(f"  Cannot find patient_number_date_time pattern in EDF filename: {edf_name}")
        return []
    
    full_pattern = match.group(1)  # patient_number_YYYYMMDD_HHMM
    
    log_print(f"  Matching pattern: {full_pattern}")
    
    # Find matching Excel files in the same directory
    edf_dir = edf_path.parent
    excel_files = []
    
    # Find all Excel files that contain the pattern (regardless of additional text)
    pattern = f"{edf_dir}/*{full_pattern}*.xlsx"
    matching_files = glob.glob(pattern)
    
    if matching_files:
        excel_files.extend(matching_files)
        log_print(f"  Found {len(matching_files)} Excel files:")
        for file in matching_files:
            log_print(f"    - {Path(file).name}")
    else:
        log_print(f"  No Excel files found with pattern: {full_pattern}")
    
    return excel_files


def get_reference_time_from_edf_metadata(edf_file):
    """Get reference time from EDF metadata start time"""
    try:
        metadata = extract_edf_metadata(edf_file)
        if not metadata:
            return None
        
        start_time_str = metadata['start_time']
        reference_seconds = parse_edf_time_to_seconds(start_time_str)
        
        log_print(f"  EDF start time: {start_time_str}")
        log_print(f"  Reference time (EDF start): {reference_seconds:.2f}s")
        return reference_seconds
        
    except Exception as e:
        log_print(f"Error getting EDF reference time: {e}")
        return None


def get_reference_time_from_events(all_events):
    """Set the earliest event time as reference for relative time calculation (DEPRECATED)"""
    if not all_events:
        return None
    
    # Find the earliest event time
    earliest_time = min(event['time_seconds'] for event in all_events)
    
    log_print(f"  Reference time (earliest event): {earliest_time:.2f}s")
    return earliest_time


def load_excel_events(excel_file, reference_time=None):
    """Load events from Excel file with relative time calculation"""
    try:
        # Try different methods to read Excel
        try:
            df = pd.read_excel(excel_file, header=None)
        except:
            df = pd.read_excel(excel_file, engine='openpyxl', header=None)
        
        # Remove first empty row
        if len(df) > 0 and df.iloc[0].isna().all():
            df = df.iloc[1:].reset_index(drop=True)
        
        log_print(f"Excel data shape: {df.shape}")
        log_print(f"All data:")
        for i in range(len(df)):
            log_print(f"  Row {i}: {df.iloc[i].tolist()}")
        
        # Extract events (time in column 2, event type in column 3)
        events = []
        for i in range(len(df)):
            time_str = str(df.iloc[i, 2]).strip()  # Column C (index 2)
            event_type = str(df.iloc[i, 3]).strip()  # Column D (index 3)
            
            log_print(f"  Processing row {i}: time='{time_str}', event='{event_type}'")
            
            # Check conditions
            has_time = ':' in time_str
            valid_event = event_type not in ['nan', 'None', ''] and str(event_type).strip() != ''
            
            if not has_time:
                log_print(f"    -> Skipped: No time format (:)")
            elif not valid_event:
                log_print(f"    -> Skipped: Invalid event type")
            
            if has_time and valid_event:
                time_seconds = parse_time_to_seconds(time_str)
                
                # Always use EDF reference time for calculation, ignore Excel relative time
                if reference_time is not None:
                    relative_time = time_seconds - reference_time
                    events.append({
                        'time': time_str,
                        'time_seconds': time_seconds,
                        'relative_time': relative_time,
                        'event_type': event_type,
                        'row_index': i  # Store original row index for status update
                    })
                    log_print(f"    -> Added event: {event_type} at {time_str} (absolute: {time_seconds:.2f}s, relative: {relative_time:.2f}s) [EDF-based]")
                else:
                    events.append({
                        'time': time_str,
                        'time_seconds': time_seconds,
                        'event_type': event_type,
                        'row_index': i  # Store original row index for status update
                    })
                    log_print(f"    -> Added event: {event_type} at {time_str} ({time_seconds:.2f}s)")
        
        log_print(f"Loaded {len(events)} events")
        for i, event in enumerate(events[:5]):  # Show first 5 events
            if 'relative_time' in event:
                log_print(f"  Event {i+1}: {event['event_type']} at {event['time']} (absolute: {event['time_seconds']:.2f}s, relative: {event['relative_time']:.2f}s)")
            else:
                log_print(f"  Event {i+1}: {event['event_type']} at {event['time']} ({event['time_seconds']:.2f}s)")
        
        return events
        
    except Exception as e:
        log_print(f"Error loading Excel file: {e}")
        return []


def update_excel_event_status(excel_file, event_status_info):
    """Update Excel file with event inclusion status"""
    try:
        import openpyxl
        
        # Load Excel file
        wb = openpyxl.load_workbook(excel_file)
        ws = wb.active
        
        # Ensure we have enough columns (add column F for status if needed)
        max_col = ws.max_column
        if max_col < 6:  # Need column F (index 6)
            for row in range(1, ws.max_row + 1):
                ws.cell(row=row, column=6, value="")
        
        # Update status for each event
        for event_info in event_status_info:
            row_idx = event_info['row_index'] + 1  # Excel is 1-indexed
            status = event_info['status']
            ws.cell(row=row_idx, column=6, value=status)
        
        # Save the file
        wb.save(excel_file)
        log_print(f"  Updated Excel file with event status: {excel_file}")
        
    except Exception as e:
        log_print(f"  Error updating Excel file status: {e}")


def convert_edf_to_edfplus(edf_file, output_file=None):
    """Convert EDF file with Excel events to EDF+ format"""
    
    log_print(f"=== EDF to EDF+ Conversion ===")
    log_print(f"EDF file: {edf_file}")
    
    try:
        # 1. Find matching Excel files
        log_print("\n1. Finding matching Excel files...")
        excel_files = find_matching_excel_files(edf_file)
        
        if not excel_files:
            log_print("  No matching Excel files found")
            return False
        
        # 2. Load EDF file with improved duration handling
        log_print("\n2. Loading EDF file...")
        
        # Read EDF header to get correct duration
        with open(edf_file, 'rb') as f:
            header = f.read(256)
            n_records = int(header[236:244].decode('ascii').strip())
            record_duration = float(header[244:252].decode('ascii').strip())
            header_duration = n_records * record_duration
        
        log_print(f"  EDF header duration: {header_duration:.2f} seconds")
        log_print(f"  EDF records: {n_records}, record duration: {record_duration:.2f}s")
        
        # Load with MNE but with better parameters
        try:
            raw = mne.io.read_raw_edf(edf_file, preload=True, verbose=False, stim_channel=None)
        except:
            # Fallback: try with different parameters
            raw = mne.io.read_raw_edf(edf_file, preload=True, verbose=False)
        
        mne_duration = raw.times[-1]
        log_print(f"  MNE reported duration: {mne_duration:.2f} seconds")
        
        # Check for duration mismatch and fix it
        duration_diff = abs(mne_duration - header_duration)
        log_print(f"  Duration difference: {duration_diff:.2f} seconds")
        
        # Store original duration for event filtering
        original_duration = mne_duration
        zero_padding_start = None
        
        if duration_diff > 0.01:  # Very sensitive threshold (0.01 seconds)
            log_print(f"  Duration mismatch detected! Fixing data...")
            target_samples = int(header_duration * raw.info['sfreq'])
            current_samples = raw.get_data().shape[1]
            
            log_print(f"  Target samples: {target_samples}, Current samples: {current_samples}")
            
            if target_samples > current_samples:
                # Pad with zeros instead of repeating last sample
                data = raw.get_data()
                padding_samples = target_samples - current_samples
                log_print(f"  Padding with {padding_samples} samples ({padding_samples/raw.info['sfreq']:.2f}s)")
                
                # Record where zero padding starts
                zero_padding_start = original_duration
                log_print(f"  Zero padding starts at: {zero_padding_start:.2f}s")
                
                # Create zero padding
                zero_padding = np.zeros((data.shape[0], padding_samples))
                padded_data = np.concatenate([data, zero_padding], axis=1)
                raw = mne.io.RawArray(padded_data, raw.info, verbose=False)
                log_print(f"  Data padded to: {raw.times[-1]:.2f} seconds")
            elif target_samples < current_samples:
                # Truncate if too long
                log_print(f"  Truncating {current_samples - target_samples} samples")
                data = raw.get_data()
                truncated_data = data[:, :target_samples]
                raw = mne.io.RawArray(truncated_data, raw.info, verbose=False)
                log_print(f"  Data truncated to: {raw.times[-1]:.2f} seconds")
        
        actual_duration = raw.times[-1]
        log_print(f"  EDF loaded: {raw.info['nchan']} channels, {actual_duration:.2f} seconds")
        log_print(f"  Sampling rate: {raw.info['sfreq']} Hz")
        log_print(f"  Final duration matches header: {abs(actual_duration - header_duration) < 0.1}")
        
        # 3. Get reference time from EDF metadata
        log_print("\n3. Getting reference time from EDF metadata...")
        reference_time = get_reference_time_from_edf_metadata(edf_file)
        
        # 4. Load events from all Excel files with EDF reference time
        log_print("\n4. Loading events from Excel files...")
        all_events = []
        excel_event_mapping = {}  # Map events to their Excel files
        
        for excel_file in excel_files:
            log_print(f"\n  Processing: {excel_file}")
            events = load_excel_events(excel_file, reference_time=reference_time)
            all_events.extend(events)
            
            # Store mapping for status updates
            for event in events:
                excel_event_mapping[len(all_events) - len(events) + events.index(event)] = excel_file
        
        if not all_events:
            log_print("  No events found, saving EDF without events")
            if output_file:
                mne.export.export_raw(output_file, raw, fmt='edf', overwrite=True)
                log_print(f"  Saved: {output_file}")
            return True
        
        # 5. Process events
        log_print(f"\n5. Processing {len(all_events)} events...")
        
        # Sort events by time
        all_events.sort(key=lambda x: x.get('relative_time', x['time_seconds']))
        
        # Debug: Show EDF duration vs last event timing
        if all_events:
            last_event = all_events[-1]
            last_relative_time = last_event.get('relative_time', last_event['time_seconds'])
            log_print(f"\n  === Duration Debug Info ===")
            log_print(f"  EDF actual duration: {actual_duration:.2f} seconds")
            log_print(f"  Last event relative time: {last_relative_time:.2f} seconds")
            log_print(f"  Time difference: {actual_duration - last_relative_time:.2f} seconds")
            log_print(f"  Last event within range: {0 <= last_relative_time <= actual_duration}")
            log_print(f"  Last event details: {last_event['event_type']} at {last_event.get('time', 'N/A')}")
            log_print(f"  =========================\n")
        
        # Convert to relative times and create annotations
        onsets = []
        durations = []
        descriptions = []
        event_status = []  # Track which events were included/excluded and why
        
        for i, event in enumerate(all_events):
            # Use relative_time if available, otherwise calculate from first event
            if 'relative_time' in event:
                relative_time = event['relative_time']
            else:
                # Fallback: use first event as reference
                if not onsets:
                    ref_time = event['time_seconds']
                    relative_time = 0.0
                else:
                    relative_time = event['time_seconds'] - ref_time
            
            # Debug: Show detailed info for last few events
            if i >= len(all_events) - 3:  # Show last 3 events
                log_print(f"  Event {i+1}/{len(all_events)}: {event['event_type']}")
                log_print(f"    Relative time: {relative_time:.2f}s")
                log_print(f"    Within range (0-{actual_duration:.2f}s): {0 <= relative_time <= actual_duration}")
                log_print(f"    Time to EDF end: {actual_duration - relative_time:.2f}s")
            
            # Check if event is in zero padding area
            in_zero_padding = False
            if zero_padding_start is not None and relative_time >= zero_padding_start:
                in_zero_padding = True
                log_print(f"  Skipped: {event['event_type']} at {relative_time:.2f}s (in zero padding area: {zero_padding_start:.2f}s+)")
                event_status.append(f"EXCLUDED_ZERO_PADDING")
                continue
            
            # Allow events that are very close to the end (within 0.1 seconds)
            tolerance = 0.1
            if relative_time >= 0 and relative_time <= actual_duration + tolerance:
                # If event is slightly beyond duration, clamp it to the end
                if relative_time > actual_duration:
                    relative_time = actual_duration
                    log_print(f"  Adjusted: {event['event_type']} clamped to EDF end ({actual_duration:.2f}s)")
                
                onsets.append(relative_time)
                durations.append(0.0)  # Point events
                descriptions.append(event['event_type'])
                event_status.append(f"INCLUDED")
                if i < len(all_events) - 3:  # Don't duplicate debug info
                    log_print(f"  Event: {event['event_type']} at {relative_time:.2f}s")
            else:
                log_print(f"  Skipped: {event['event_type']} at {relative_time:.2f}s (outside EDF range: 0-{actual_duration:.2f}s)")
                event_status.append(f"EXCLUDED_OUT_OF_RANGE")
        
        # 6. Add annotations to raw data
        log_print("\n6. Adding events to EDF...")
        annotations = mne.Annotations(
            onset=onsets,
            duration=durations,
            description=descriptions
        )
        raw.set_annotations(annotations)
        log_print(f"  Added {len(onsets)} events to EDF")
        
        # 6.5. Update Excel files with event status
        log_print("\n6.5. Updating Excel files with event status...")
        excel_status_updates = {}  # Group status updates by Excel file
        
        for i, status in enumerate(event_status):
            if i < len(all_events):
                event = all_events[i]
                excel_file = excel_event_mapping.get(i)
                if excel_file:
                    if excel_file not in excel_status_updates:
                        excel_status_updates[excel_file] = []
                    excel_status_updates[excel_file].append({
                        'row_index': event['row_index'],
                        'status': status
                    })
        
        # Update each Excel file
        for excel_file, status_updates in excel_status_updates.items():
            update_excel_event_status(excel_file, status_updates)
        
        # 7. Generate correct filename and save as EDF+
        log_print("\n7. Generating correct filename and saving EDF+ file...")
        
        # Generate correct filename based on EDF metadata
        correct_filename = generate_correct_filename(edf_file)
        if not correct_filename:
            log_print("  Warning: Could not generate correct filename, using original filename")
            correct_filename = os.path.basename(edf_file)
        
        if output_file is None:
            # Rename original file to backup with final filename info
            edf_path = Path(edf_file)
            original_stem = edf_path.stem  # original file name (without extension)
            final_stem = correct_filename.replace('.edf', '')  # final file name (without extension)
            
            # backup file name: original file name_backup_final file name.edf
            backup_file = edf_path.parent / f"{original_stem}_backup_{final_stem}.edf"
            
            # Rename original file to backup
            import shutil
            shutil.move(edf_file, backup_file)
            log_print(f"  Original file renamed to backup: {backup_file.name}")
            
            # Use correct filename for EDF+ output
            output_file = edf_path.parent / correct_filename
        
        # Save using MNE's built-in EDF+ export functionality
        # Verify duration before saving
        final_duration = raw.times[-1]
        log_print(f"  Final duration before saving: {final_duration:.2f} seconds")
        
        mne.export.export_raw(output_file, raw, fmt='edf', overwrite=True)
        
        # Verify the saved file duration
        try:
            saved_raw = mne.io.read_raw_edf(output_file, preload=False, verbose=False)
            saved_duration = saved_raw.times[-1]
            log_print(f"  Saved file duration: {saved_duration:.2f} seconds")
            
            if abs(saved_duration - final_duration) > 0.1:
                log_print(f"  ⚠️  WARNING: Duration mismatch after saving!")
                log_print(f"     Original: {final_duration:.2f}s, Saved: {saved_duration:.2f}s")
            else:
                log_print(f"  ✓ Duration preserved correctly")
        except Exception as e:
            log_print(f"  Could not verify saved file duration: {e}")
        
        log_print(f"  ✓ EDF+ file saved: {output_file}")
        log_print(f"  ✓ Filename based on EDF metadata: {correct_filename}")
        
        return True
        
    except Exception as e:
        log_print(f"Error during conversion: {e}")
        return False


def run_relative_time_processing():
    """Run relative.py to add relative time to Excel files"""
    log_print("=== Running relative time processing ===")
    try:
        # Import and run the relative time processing
        import subprocess
        import sys
        
        result = subprocess.run([sys.executable, "relative.py"], 
                              capture_output=True, text=True, cwd=".")
        
        if result.returncode == 0:
            log_print("✓ Relative time processing completed successfully")
            log_print(result.stdout)
        else:
            log_print("✗ Relative time processing failed")
            log_print("STDOUT:", result.stdout)
            log_print("STDERR:", result.stderr)
            return False
            
    except Exception as e:
        log_print(f"Error running relative time processing: {e}")
        return False
    
    return True


def process_all_edf_files(base_dir="."):
    """Process all EDF files in the directory and subdirectories"""
    base_path = Path(base_dir)
    
    # First, run relative time processing
    log_print("Step 1: Adding relative time to Excel files...")
    if not run_relative_time_processing():
        log_print("Failed to process relative time. Continuing with EDF conversion...")
    
    log_print("\nStep 2: Converting EDF files to EDF+...")
    
    # Find all EDF files recursively
    edf_files = list(base_path.rglob("*.edf"))
    
    # Filter out backup files and already processed files
    edf_files = [f for f in edf_files if "_backup" not in f.name and "_with_events" not in f.name]
    
    if not edf_files:
        log_print("No EDF files found in the directory")
        return
    
    log_print(f"=== Found {len(edf_files)} EDF files to process ===")
    for i, edf_file in enumerate(edf_files, 1):
        log_print(f"{i}. {edf_file}")
    
    # Process each EDF file
    success_count = 0
    failed_files = []
    
    for i, edf_file in enumerate(edf_files, 1):
        log_print(f"\n{'='*60}")
        log_print(f"Processing {i}/{len(edf_files)}: {edf_file.name}")
        log_print(f"{'='*60}")
        
        try:
            success = convert_edf_to_edfplus(edf_file)
            if success:
                success_count += 1
                log_print(f"✓ Success: {edf_file.name}")
            else:
                failed_files.append(edf_file)
                log_print(f"✗ Failed: {edf_file.name}")
        except Exception as e:
            failed_files.append(edf_file)
            log_print(f"✗ Error processing {edf_file.name}: {e}")
    
    # Summary
    log_print(f"\n{'='*60}")
    log_print(f"=== PROCESSING COMPLETE ===")
    log_print(f"Total files: {len(edf_files)}")
    log_print(f"Successful: {success_count}")
    log_print(f"Failed: {len(failed_files)}")
    
    if failed_files:
        log_print(f"\nFailed files:")
        for failed_file in failed_files:
            log_print(f"  - {failed_file}")
    
    log_print(f"{'='*60}")


def main():
    """Main function"""
    import sys
    
    # Setup logging
    log_filename = setup_logging()
    log_print(f"=== EDF2EDF+ Converter Started ===")
    log_print(f"Log file: {log_filename}")
    
    if len(sys.argv) > 1:
        # If directory path provided as argument
        base_dir = sys.argv[1]
        log_print(f"Processing directory: {base_dir}")
    else:
        # Use current directory
        base_dir = "."
        log_print(f"Processing current directory: {os.getcwd()}")
    
    process_all_edf_files(base_dir)
    
    log_print(f"=== EDF2EDF+ Converter Finished ===")
    log_print(f"Log saved to: {log_filename}")


if __name__ == "__main__":
    main()
